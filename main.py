from sys import maxsize
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import logging

# Настройка логирования
logging.basicConfig(level=logging.INFO, filename='project.log', filemode='w',
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Параметры файлов
source_file_1 = 'Состав кафедры.xlsx'
source_file_2 = 'Проект нагрузки.xlsx'
target_file = 'План по нагрузке.xlsx'
header_file = 'Нагруз без лишних данных.xlsx'

# Словарь с нагрузкой для каждой должности
load_by_position = {
    'Должность': 'Допустимая нагрузка',
    'Заведующий кафедрой': 800,
    'Профессор': 800,
    'Доцент': 850,
    'Старший преподаватель': 900,
    'Ассистент': 900
}

# Ключевые слова для поиска
keywords_assistant = ['лекция', 'лекции', 'курсовая', 'курсовые', 'экзамены']
keywords_senior_lecturer = ['практика', 'практическое занятие']

# Цвет для подсветки
highlight_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

try:
    df_header = pd.read_excel(header_file, header=None)
    header = df_header.values.tolist()
    logging.info(f"Заголовки успешно прочитаны из файла '{header_file}'.")
except Exception as e:
    logging.error(f"Ошибка при обработке файла '{header_file}': {e}")
    header = []

try:
    df_orders = pd.read_excel(source_file_1, header=None, usecols=[1, 3])
    df_orders = df_orders.dropna().drop_duplicates()
    df_orders = df_orders[~((df_orders[1] == 'Федосенко Юрий Семенович') & (df_orders[3] == 'Профессор'))]
    df_orders = df_orders[~(df_orders[1] == 'Вакантно')]
    df_orders = df_orders.copy()
    df_orders['Нагрузка'] = df_orders[3].map(load_by_position)
    rows_orders = df_orders.values.tolist()
    logging.info(f"Данные успешно прочитаны из файла '{source_file_1}'.")
except Exception as e:
    logging.error(f"Ошибка при обработке файла '{source_file_1}': {e}")
    rows_orders = []

try:
    df_nagruzka = pd.read_excel(source_file_2, header=None, skiprows=8)
    rows_nagruzka = df_nagruzka.values.tolist()
    logging.info(f"Данные успешно прочитаны из файла '{source_file_2}'.")
except Exception as e:
    logging.error(f"Ошибка при обработке файла '{source_file_2}': {e}")
    rows_nagruzka = []

combined_rows = []
if header:
    combined_rows.append(header[0])

for nagruzka_row in rows_nagruzka:
    combined_rows.append(nagruzka_row)
    combined_rows.extend(rows_orders)
    combined_rows.append('\n')

combined_df = pd.DataFrame(combined_rows)

combined_df.to_excel(target_file, index=False, header=False, freeze_panes=(1, 0))
logging.info(f"Данные успешно записаны в файл '{target_file}'.")

# Подсветка ячеек в итоговом файле
try:
    workbook = load_workbook(target_file)
    sheet = workbook.active

    # Ищем столбцы, которые содержат ключевые слова в заголовке
    keyword_columns = []
    for col_idx, cell in enumerate(sheet[1], start=1):
        if isinstance(cell.value, str) and any(
                keyword in cell.value.lower() for keyword in keywords_assistant + keywords_senior_lecturer):
            keyword_columns.append(col_idx)
            logging.info(f"Столбец {col_idx} содержит ключевое слово: {cell.value}")

    # Проходим по строкам и подсвечиваем ячейки
    for row_idx in range(2, sheet.max_row + 1):  # Пропускаем заголовок
        position = sheet.cell(row=row_idx, column=2).value  # Предполагается, что "Должность" во втором столбце
        if position in ['Ассистент', 'Старший преподаватель']:
            for col_idx in keyword_columns:
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.fill = highlight_fill
                logging.info(f"Закрашиваем ячейку: строка {row_idx}, столбец {col_idx}")

    workbook.save(target_file)
    logging.info(f"Подсветка завершена. Файл '{target_file}' обновлен.")
except Exception as e:
    logging.error(f"Ошибка при подсветке ключевых слов: {e}")
